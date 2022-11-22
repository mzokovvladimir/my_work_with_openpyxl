from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

DATA = {
    'Ivan': {
        'physics': 91,
        'math': 81,
        'science': 72
    },
    'Bogdan': {
        'physics': 81,
        'math': 83,
        'science': 82
    },
    'Danil': {
        'physics': 76,
        'math': 77,
        'science': 75
    },
    'Nina': {
        'physics': 91,
        'math': 93,
        'science': 98
    },
    'Dasha': {
        'physics': 71,
        'math': 73,
        'science': 75
    },
    'Alex': {
        'physics': 80,
        'math': 87,
        'science': 85
    },
    'Irina': {
        'physics': 71,
        'math': 69,
        'science': 89
    },
    'Ilona': {
        'physics': 72,
        'math': 74,
        'science': 78
    }
}

wb = Workbook()
ws = wb.active
ws.title = 'grades'

headings = ['name'] + list(DATA['Ivan'].keys())
ws.append(headings)

for student in DATA:
    grades = list(DATA[student].values())
    ws.append([student] + grades)

ws['A10'].value = 'Average1:'
for col in range(2, len(DATA['Ivan']) + 2):
    char = get_column_letter(col)
    ws[char + '10'] = f"=SUM({char + '2'}:{char + '9'}) / {len(DATA)}"

ws['A11'].value = 'Average2:'
for col in range(2, len(DATA['Ivan']) + 2):
    char = get_column_letter(col)
    ws[char + '11'] = f"=ROUND(AVERAGE({char + '2'}:{char + '9'}), 2)"

for col in range(1, 6):
    ws[get_column_letter(col) + '1'].font = Font(bold=True, color='FF883355')

wb.save('grades.xlsx')
