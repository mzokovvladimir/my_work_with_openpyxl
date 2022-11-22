from openpyxl import load_workbook

wb = load_workbook('test.xlsx')
ws = wb.active

ws.merge_cells("A1:D1")
ws.unmerge_cells("A1:D1")

wb.save('test1.xlsx')
