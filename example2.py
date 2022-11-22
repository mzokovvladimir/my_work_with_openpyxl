from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('test.xlsx')
ws = wb.active

for row in range(1, 17):
    for col in range(1, 6):
        char = get_column_letter(col)
        ws[char + str(row)] = char + str(row)

wb.save('test1.xlsx')
